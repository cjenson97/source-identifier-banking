"""Tests for source_identifier_banking.excel_importer."""

from __future__ import annotations

import pytest
import yaml
import openpyxl

from source_identifier_banking.excel_importer import read_excel_sources, import_sources


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_xlsx(tmp_path, headers: list, rows: list[list], filename: str = "sources.xlsx") -> str:
    """Write a simple .xlsx file and return its path as a string."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    path = tmp_path / filename
    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# read_excel_sources
# ---------------------------------------------------------------------------

class TestReadExcelSources:
    def test_basic_url_column(self, tmp_path):
        path = _make_xlsx(
            tmp_path,
            headers=["url"],
            rows=[["https://example.com"], ["https://another.org"]],
        )
        sources = read_excel_sources(path)
        assert len(sources) == 2
        assert sources[0]["domain"] == "example.com"
        assert sources[1]["domain"] == "another.org"

    def test_url_column_aliases(self, tmp_path):
        for col in ("URL", "Link", "homepage", "DOMAIN", "Website"):
            path = _make_xlsx(tmp_path, headers=[col], rows=[["https://test.io"]])
            sources = read_excel_sources(path)
            assert len(sources) == 1, f"alias '{col}' not recognised"
            assert sources[0]["domain"] == "test.io"

    def test_optional_columns_mapped(self, tmp_path):
        path = _make_xlsx(
            tmp_path,
            headers=["url", "name", "jurisdiction", "source_type", "feed_url"],
            rows=[["https://regulator.gov", "My Regulator", "EU", "regulator", "https://regulator.gov/feed"]],
        )
        sources = read_excel_sources(path)
        assert len(sources) == 1
        s = sources[0]
        assert s["name"] == "My Regulator"
        assert s["jurisdiction"] == "EU"
        assert s["source_type"] == "regulator"
        assert s["feed_url"] == "https://regulator.gov/feed"

    def test_domain_extracted_from_full_url(self, tmp_path):
        path = _make_xlsx(
            tmp_path,
            headers=["url"],
            rows=[["https://www.example.com/some/path?q=1"]],
        )
        sources = read_excel_sources(path)
        assert sources[0]["domain"] == "example.com"

    def test_bare_domain_in_url_column(self, tmp_path):
        path = _make_xlsx(tmp_path, headers=["url"], rows=[["example.com"]])
        sources = read_excel_sources(path)
        assert sources[0]["domain"] == "example.com"

    def test_missing_url_column_raises(self, tmp_path):
        path = _make_xlsx(tmp_path, headers=["source_name", "notes"], rows=[["Foo", "bar"]])
        with pytest.raises(ValueError, match="No URL/domain column found"):
            read_excel_sources(path)

    def test_file_not_found_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            read_excel_sources(str(tmp_path / "nonexistent.xlsx"))

    def test_empty_rows_skipped(self, tmp_path):
        path = _make_xlsx(
            tmp_path,
            headers=["url"],
            rows=[["https://valid.com"], [None], [""], ["https://another.com"]],
        )
        sources = read_excel_sources(path)
        assert len(sources) == 2

    def test_defaults_when_optional_columns_absent(self, tmp_path):
        path = _make_xlsx(tmp_path, headers=["url"], rows=[["https://minimal.io"]])
        sources = read_excel_sources(path)
        assert sources[0]["jurisdiction"] == "Unknown"
        assert sources[0]["source_type"] == "unknown"
        assert sources[0]["feed_url"] is None
        assert sources[0]["name"] == "minimal.io"

    def test_named_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Data")
        ws.append(["url"])
        ws.append(["https://sheettest.com"])
        path = tmp_path / "multi.xlsx"
        wb.save(str(path))

        sources = read_excel_sources(str(path), sheet="Data")
        assert len(sources) == 1
        assert sources[0]["domain"] == "sheettest.com"


# ---------------------------------------------------------------------------
# import_sources
# ---------------------------------------------------------------------------

class TestImportSources:
    def test_creates_yaml_from_xlsx(self, tmp_path):
        xlsx = _make_xlsx(tmp_path, headers=["url", "name"], rows=[["https://new.org", "New Org"]])
        yaml_out = str(tmp_path / "existing.yaml")

        added = import_sources(xlsx, yaml_out)
        assert added == 1

        data = yaml.safe_load(open(yaml_out).read())
        assert len(data["sources"]) == 1
        assert data["sources"][0]["domain"] == "new.org"

    def test_merges_without_duplicates(self, tmp_path):
        existing_yaml = tmp_path / "existing.yaml"
        existing_yaml.write_text(
            "sources:\n  - name: Old Source\n    domain: old.org\n"
        )
        xlsx = _make_xlsx(
            tmp_path,
            headers=["url"],
            rows=[["https://new.org"], ["https://old.org"]],  # old.org is a duplicate
        )

        added = import_sources(str(xlsx), str(existing_yaml))
        assert added == 1  # only new.org added

        data = yaml.safe_load(existing_yaml.read_text())
        domains = {s["domain"] for s in data["sources"]}
        assert domains == {"old.org", "new.org"}

    def test_replace_overwrites_existing(self, tmp_path):
        existing_yaml = tmp_path / "existing.yaml"
        existing_yaml.write_text(
            "sources:\n  - name: Old Source\n    domain: old.org\n"
        )
        xlsx = _make_xlsx(tmp_path, headers=["url"], rows=[["https://new.org"]])

        added = import_sources(str(xlsx), str(existing_yaml), replace=True)
        assert added == 1

        data = yaml.safe_load(existing_yaml.read_text())
        assert len(data["sources"]) == 1
        assert data["sources"][0]["domain"] == "new.org"

    def test_creates_parent_directories(self, tmp_path):
        xlsx = _make_xlsx(tmp_path, headers=["url"], rows=[["https://x.com"]])
        yaml_out = str(tmp_path / "nested" / "deep" / "out.yaml")

        import_sources(xlsx, yaml_out)
        assert yaml.safe_load(open(yaml_out).read())["sources"][0]["domain"] == "x.com"

    def test_returns_zero_when_all_duplicates(self, tmp_path):
        existing_yaml = tmp_path / "existing.yaml"
        existing_yaml.write_text("sources:\n  - domain: dup.com\n")
        xlsx = _make_xlsx(tmp_path, headers=["url"], rows=[["https://dup.com"]])

        added = import_sources(str(xlsx), str(existing_yaml))
        assert added == 0
